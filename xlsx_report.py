#!/usr/bin/env python3.11
# -*- coding: utf-8 -*-

# script header
# Author:           [Hugo Berra Salazar, ]
# Creation date:    [04/07/2023]
# Description:      [Brief description of the purpose of the script]

# import necessary modules
import os
import urllib.request
import xml.etree.ElementTree as ET

from datetime           import datetime
from openpyxl.styles    import Alignment
from openpyxl           import load_workbook, Workbook
from openpyxl.styles    import Color, Font, PatternFill

# Define the ElementPath queries
version_query               = 'Version'

def get_dir_path_data(option):
    dirs = {
        'AYUDAS'        : '/Emisor/Ayudas/',
        'INGRESO'       : '/Emisor/Ingresos/',
        'GASTOE'        : '/Emisor/Gastos/',
        'NOMINA'        : '/Emisor/Nomina/',
        'DES_BON_DEV'   : '/Receptor/Descuento_Bonificaciones_Devoluciones/',
        'GASTOR'        : '/Receptor/Gastos/',
        'PAGOS'         : '/Receptor/Pagos/'
    }
    return dirs.get(option)

def filenames(dir):
    for root, dirs, files in os.walk(dir):
        for file in files:
            if file.endswith(".xml"):
                yield os.path.join(root, file)

def rename_xlsx_headers(path):
    wb = load_workbook(path)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    count = {}

    for i, header in enumerate(headers):
        if header in count:
            count[header] += 1
            headers[i] = f"{header} receptor"
        else:
            count[header] = 0

    for col_num, name in enumerate(headers, 1):
         cell = ws.cell(row=1, column=col_num, value=name)

    wb.save(path)
    print(f'{path}: xlsx namespaces modified successfully')

def get_cfdi_version(version):
    if   version == '4.0':
        emisor_query    = '{http://www.sat.gob.mx/cfd/4}Emisor'
        receptor_query  = '{http://www.sat.gob.mx/cfd/4}Receptor'
    elif version == '3.3':
        emisor_query    = '{http://www.sat.gob.mx/cfd/3}Emisor'
        receptor_query  = '{http://www.sat.gob.mx/cfd/3}Receptor'
    return emisor_query, receptor_query

def del_empty_columns(path):
    # Load the workbook
    wb = load_workbook(path)
    ws = wb.active

    # Find the maximum number of rows and columns
    max_row = ws.max_row
    max_col = ws.max_column

    # Loop through each column
    for col in range(1, max_col + 1):
        # Check if all cells in column are empty starting from row 2
        empty = True
        for row in range(2, max_row + 1):
            if ws.cell(row=row, column=col).value != None:
                empty = False
                break
            
        # If all cells are empty, delete the column
        if empty:
            ws.delete_cols(col, 1)
    # Save the workbook
    wb.save(path)
    print(f'{path} xlsx deleted columns successfully.')

def create_xlsx(rfc, option):
    try:
        # Use set() method to save attribute names reduces the amount of memory needed
        url_cfdi40  = 'http://www.sat.gob.mx/sitio_internet/cfd/4/cfdv40.xsd'
        url_cfdi33  = 'http://www.sat.gob.mx/sitio_internet/cfd/3/cfdv33.xsd'
        cfdi40      = []
        cfdi33      = []

        # Reads both CFDI standards and extracts the attribute names from each CFDI version. 
        # These names are stored in two separate lists
        with urllib.request.urlopen(url_cfdi40) as response:
            tree = ET.parse(response)
            root = tree.getroot()

            for nodo in root.iter():
                if 'name' in nodo.attrib:
                    cfdi40.append(nodo.attrib['name'])

        with urllib.request.urlopen(url_cfdi33) as response:
            tree = ET.parse(response)
            root = tree.getroot()

            for nodo in root.iter():
                if 'name' in nodo.attrib:
                    cfdi33.append(nodo.attrib['name'])

        # filter attributes only present in 4.0            
        attr_40_list = [attr for attr in cfdi40 if attr not in cfdi33]

        # Gets the current date and time
        fecha_actual = datetime.now().strftime('%m%d%Y-%H%M%S')

        # In the rfc folder and its corresponding subfolder, using the name of the previously created file.
        try:
            os.makedirs(os.path.join(rfc, 'xlsx_report', option))
        except FileExistsError:
            pass

        filename    = f"{option}-{fecha_actual}.xlsx"            
        xlsx_path   = os.path.join(rfc, 'xlsx_report', option, filename)
        
        # Create an Excel file and format the header
        wb = Workbook()
        ws = wb.active
        ws.title = f"{option}_{fecha_actual}"
        header_font = Font(color='FFFFFF')
        header_fill = PatternFill(start_color='730707', end_color='730707', fill_type='solid')

    # write headers for version 3.3
        for col_num, name in enumerate(cfdi33, 1):
            cell = ws.cell(row=1, column=col_num, value=name)
            cell.font = header_font
            cell.fill = header_fill

        # write headers for version 4.0
        for col_num, name in enumerate(attr_40_list, len(cfdi33) + 1):
            cell = ws.cell(row=1, column=col_num, value=name)
            cell.font = header_font
            cell.fill = header_fill

        # Save workbook
        wb.save(xlsx_path)
        print(f'{xlsx_path}: xlsx created successfully')
        return xlsx_path

    except Exception as e:
        print(f"Error: {e}")

def cfdi_to_xlsx(rfc, option):
    # 
    path = create_xlsx(rfc, option)
    rename_xlsx_headers(path)

    # 
    wb      = load_workbook(path)
    ws      = wb.active
    headers = [cell.value for cell in ws[1]]

    for filename in filenames(f"{rfc}{get_dir_path_data(option)}"):
        try:
            tree = ET.parse(filename)
            root = tree.getroot()

            # Get the cfdi version
            version                       = root.get(version_query)
            emisor_query, receptor_query  = get_cfdi_version(version)
            
            emisor      = root.find(emisor_query)
            receptor    = root.find(receptor_query)

            if emisor is not None and receptor is not None:
                row_data = {}
                for nodo in root.iter():
                    if nodo.attrib:
                        for key, value in nodo.attrib.items():
                            if key in headers:
                                row_data['Rfc']             = str(emisor.get('Rfc'))
                                row_data['Rfc receptor']    = str(receptor.get('Rfc'))
                                row_data['Nombre']          = str(emisor.get('Nombre'))
                                row_data['Nombre receptor'] = str(receptor.get('Nombre'))
                                row_data[key]               = value
                if row_data:
                    row_num = ws.max_row + 1
                    for col_num, header in enumerate(headers, 1):
                        if header in row_data:
                            if header == 'Total' or header == 'SubTotal' or header == 'TotalImpuestosRetenidos' or header == 'Descuento' or header == 'Importe' or header == 'ValorUnitario' or header == 'Base' or header == 'TasaOCuota':
                                ws.cell(row=row_num, column=col_num, value=float(row_data[header]))
                            else:
                                ws.cell(row=row_num, column=col_num, value=row_data[header]) 
            #saving the values that correspond to each header.
            wb.save(path)
            print(f'{filename} xlsx filled successfully.')
        except ET.ParseError:
            print(f"{filename} could not be parsed.")
        except Exception as e:
            print(f"{filename} could not be processed due to an error: {e}.")
        
    del_empty_columns(path)